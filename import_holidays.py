from openpyxl import load_workbook
from configobj import ConfigObj
from sqlalchemy.sql import text
from sqlalchemy import create_engine

config = ConfigObj(".env")
database_uri = "mysql+mysqldb://%(database_user)s:%(database_password)s@%(database_host)s/%(database_name)s" % config
engine = create_engine(database_uri)
connection = engine.connect()
connection.execute(text("delete from holidays"))

def add_holiday(date):
    connection = engine.connect()
    print(date)
    connection.execute(text("insert into holidays (week_number, weekday, year) values (weekofyear(:date), weekday(:date), year(:date))").bindparams(date=date))


wb = load_workbook('holidays-jul2016-jun2017.xlsx')
ws = wb.active
for cell in ws['A']:
    if (cell.is_date):
        if (cell.value != None):
            add_holiday(cell.value)
